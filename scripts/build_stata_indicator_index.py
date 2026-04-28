"""
Step 1 of the DHS Stata indicator pipeline.

DATA FLOW
  IndicatorList.xlsx  (DHS Code Share Project, one sheet per chapter)
      ↓  this script
  references/dhs_stata_indicators.json  (1,225 indicators, fields: stata_var,
      label, chapter, do_file, dhs_file, notes)
      ↓  extract_stata_dhs_vars.py (Step 2)
  same JSON enriched with dhs_variables + ipums_variables per indicator

WHY THIS EXISTS
  The DHS Code Share Project (github.com/DHSProgram/DHS-Indicators-Stata)
  publishes Stata .do files that implement every standard DHS indicator.
  The accompanying IndicatorList.xlsx catalogues all 1,225 indicators with
  their Stata variable name, human-readable label, and source .do file.
  This script converts that xlsx into a JSON index so Step 2 can look up
  which .do file to parse for each indicator.

XLSX STRUCTURE
  18 sheets, one per chapter (e.g. "Chap02_PH", "Chap10_CH").
  Six columns per row: Chapter, Do file, Var name, Var Label, File, Notes.
  Chapter and Do file cells are only populated on section-header rows;
  they carry forward implicitly for all indicator rows beneath them.
  Some rows are structural headers (e.g. "Main file: KR dataset") rather
  than actual indicators and must be skipped.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

XLSX_PATH = Path("/Users/arome/Desktop/IndicatorList (1).xlsx")
OUT_PATH = Path(__file__).parent.parent / "references" / "dhs_stata_indicators.json"

# Matches the leading DHS file code (HR, IR, KR, PR, MR, BR) in the File/Notes column.
# Stops at a word boundary so "IR or KR" extracts "IR", not "IRON".
_FILE_CODE_RE = re.compile(r"^([A-Z]{2,3})\b")


def parse_dhs_file(file_val, notes_val):
    """Return primary DHS file code (e.g. 'KR') from whichever column has it, or None."""
    for candidate in (file_val, notes_val):
        if not candidate or not isinstance(candidate, str):
            continue
        m = _FILE_CODE_RE.match(candidate.strip())
        if m:
            return m.group(1)
    return None


def clean_label(raw):
    if not raw or not isinstance(raw, str):
        return None
    return raw.strip().strip('"').strip("'").strip()


def parse_sheet(ws):
    """Yield one indicator dict per data row in the worksheet.

    Chapter and do_file are carried forward from the last non-None cell
    because the xlsx only repeats them on the first row of each section.
    Structural header rows are recognised by their Var name cell content
    and skipped before the space-strip so the check works on the raw value.
    """
    current_chapter = None
    current_do_file = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        chapter_col, dofile_col, var_col, label_col, file_col, notes_col = row[:6]

        if chapter_col and isinstance(chapter_col, str):
            current_chapter = chapter_col.strip()

        if dofile_col and isinstance(dofile_col, str):
            current_do_file = dofile_col.strip()

        if not var_col or not isinstance(var_col, str):
            continue
        var_raw = var_col.strip()
        if not var_raw:
            continue

        # Skip section-header rows like "Main file: KR dataset" and "Do file: CH_VAC.do".
        # These must be checked on var_raw (before space removal) because removing spaces
        # turns "Main file" into "Mainfile", breaking the startswith check.
        if var_raw.startswith("Main file") or var_raw.startswith("Do file"):
            continue
        if var_raw == "Var name":  # column-header row
            continue

        # A handful of xlsx cells have internal spaces, e.g. "ph_cook_ clean".
        # Strip them so the var name matches the actual Stata identifier.
        var_stripped = var_raw.replace(" ", "")

        yield {
            "stata_var": var_stripped,
            "label": clean_label(label_col),
            "chapter": current_chapter,
            "do_file": current_do_file,
            "dhs_file": parse_dhs_file(file_col, notes_col),
            "notes": notes_col.strip() if isinstance(notes_col, str) else notes_col,
        }


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    indicators = []
    chapter_counts = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_indicators = list(parse_sheet(ws))
        indicators.extend(sheet_indicators)

        for ind in sheet_indicators:
            ch = ind["chapter"] or sheet_name
            chapter_counts[ch] = chapter_counts.get(ch, 0) + 1

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w") as f:
        json.dump(indicators, f, indent=2)

    print(f"Total indicators: {len(indicators)}")
    print("\nPer chapter:")
    for chapter, count in sorted(chapter_counts.items(), key=lambda x: -x[1]):
        print(f"  {count:4d}  {chapter}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
