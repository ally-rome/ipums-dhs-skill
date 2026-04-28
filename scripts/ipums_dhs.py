"""
IPUMS DHS microdata CLI and importable module.

CLI usage:
    python3 scripts/ipums_dhs.py samples KE
    python3 scripts/ipums_dhs.py search "stunting"
    python3 scripts/ipums_dhs.py table --country KE --survey latest \\
        --variables HWHAZWHO --unit children --by WEALTHQ

Importable functions:
    find_samples, find_variables, submit_extract, wait_for_extract,
    download_extract, load_extract, parse_ddi_missing,
    weighted_proportion, weighted_mean, tabulate
"""

import argparse
import gzip
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Optional

import pandas as pd
import requests
import warnings
from ipumspy import readers

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Paths and constants
# ---------------------------------------------------------------------------

REFERENCES_DIR = Path(__file__).parent.parent / "references"
DATA_DIR = Path(__file__).parent.parent / "data"
AVAILABILITY_PATH = REFERENCES_DIR / "dhs_availability.json"

UNIT_SUFFIX = {
    "women": "ir",
    "men": "mr",
    "children": "kr",
    "household_members": "pr",
    "births": "br",
}

WEIGHT_DEFAULT = {
    "women": "PERWEIGHT",
    "men": "PERWEIGHTMN",
    "children": "PERWEIGHT",
    "household_members": "HHWEIGHT",
    "births": "PERWEIGHT",
}

UNIT_DESCRIPTIONS = {
    "women": "Women",
    "men": "Men",
    "children": "Children under 5",
    "household_members": "Household Members",
    "births": "Births",
}

DHS_FILE_TO_UNIT = {
    "IR": "women",
    "KR": "children",
    "PR": "household_members",
    "HR": "household_members",
    "MR": "men",
    "BR": "births",
}

# Human-readable indicator names for common anthropometric z-score variables
_INDICATOR_NAMES = {
    "HWHAZWHO": "Stunted", "HWHAZNCHS": "Stunted",
    "HWWAZWHO": "Underweight", "HWWAZNCHS": "Underweight",
    "HWWHZWHO": "Wasted", "HWWHZNCHS": "Wasted",
}

CODEBOOK_FILES = {
    "women": REFERENCES_DIR / "dhs_codebook_women.md",
    "men": REFERENCES_DIR / "dhs_codebook_men.md",
    "children": REFERENCES_DIR / "dhs_codebook_children.md",
    "household_members": REFERENCES_DIR / "dhs_codebook_household_members.md",
    "births": REFERENCES_DIR / "dhs_codebook_births.md",
}

IPUMS_API = "https://api.ipums.org/extracts"
IPUMS_PARAMS = {"collection": "dhs", "version": 2}

# Maps IPUMS 2-letter country codes (lowercase) to country names as they appear
# in references/dhs_availability.json (scraped from the IPUMS DHS variable pages).
COUNTRY_NAMES: dict[str, str] = {
    "af": "Afghanistan", "al": "Albania", "ao": "Angola", "am": "Armenia",
    "az": "Azerbaijan", "bd": "Bangladesh", "bj": "Benin", "bo": "Bolivia",
    "bf": "Burkina Faso", "bu": "Burundi", "kh": "Cambodia", "cm": "Cameroon",
    "td": "Chad", "co": "Colombia", "km": "Comoros",
    "cd": "Congo (Democratic Republic)", "cg": "Congo Brazzaville",
    "ci": "Cote d'Ivoire", "do": "Dominican Republic", "eg": "Egypt",
    "sz": "Eswatini (Swaziland)", "et": "Ethiopia", "ga": "Gabon",
    "gm": "Gambia", "gh": "Ghana", "gt": "Guatemala", "gn": "Guinea",
    "gy": "Guyana", "ht": "Haiti", "hn": "Honduras", "ia": "India",
    "id": "Indonesia", "jo": "Jordan", "kz": "Kazakhstan", "ke": "Kenya",
    "kg": "Kyrgyz Republic", "ls": "Lesotho", "lr": "Liberia",
    "mg": "Madagascar", "mw": "Malawi", "mv": "Maldives", "ml": "Mali",
    "mr": "Mauritania", "ma": "Morocco", "mz": "Mozambique", "mm": "Myanmar",
    "nm": "Namibia", "np": "Nepal", "ne": "Niger", "ng": "Nigeria",
    "pk": "Pakistan", "pg": "Papua New Guinea", "pe": "Peru", "ph": "Philippines",
    "rw": "Rwanda", "sn": "Senegal", "sl": "Sierra Leone", "za": "South Africa",
    "sd": "Sudan", "tj": "Tajikistan", "tz": "Tanzania", "tl": "Timor-Leste",
    "tg": "Togo", "tn": "Tunisia", "tr": "Turkey", "ug": "Uganda",
    "ua": "Ukraine", "uz": "Uzbekistan", "vn": "Vietnam", "ye": "Yemen",
    "zm": "Zambia", "zw": "Zimbabwe",
}

# ---------------------------------------------------------------------------
# find_samples
# ---------------------------------------------------------------------------


def find_samples(country_code: str, unit: str = None) -> list[dict]:
    """
    Return available DHS samples for a country, optionally filtered by unit.

    Args:
        country_code: Two-letter ISO country code (case-insensitive), e.g. "KE".
        unit: One of women/men/children/household_members/births, or None for all.

    Returns:
        List of dicts with keys: sample_id, country, year, unit, description.
        Sorted by year ascending.
    """
    from ipumspy import IpumsApiClient

    api_key = _require_api_key()
    client = IpumsApiClient(api_key)
    all_samples = client.get_all_sample_info("dhs")

    cc = country_code.lower()
    suffix_filter = UNIT_SUFFIX.get(unit) if unit else None

    results = []
    for sample_id, description in all_samples.items():
        if not sample_id.startswith(cc):
            continue
        if suffix_filter and not sample_id.endswith(suffix_filter):
            continue

        # Parse year and unit suffix from sample_id
        # Format: {cc}{year}{suffix}  e.g. ke2014kr
        suffix = sample_id[len(cc) + 4:]
        year_str = sample_id[len(cc):len(cc) + 4]
        unit_name = next((u for u, s in UNIT_SUFFIX.items() if s == suffix), suffix)

        results.append({
            "sample_id": sample_id,
            "country": description.split()[0],
            "year": int(year_str),
            "unit": unit_name,
            "description": description,
        })

    results.sort(key=lambda r: (r["year"], r["unit"]))
    return results


# ---------------------------------------------------------------------------
# find_variables
# ---------------------------------------------------------------------------


def find_variables(search_term: str) -> list[dict]:
    """
    Search all codebook files for variables matching a keyword (case-insensitive).

    Returns:
        List of dicts with keys: variable, label, unit, preselected.
        Ordered by unit then variable name.
    """
    pattern = re.compile(re.escape(search_term), re.IGNORECASE)
    results = []
    seen = set()  # (variable, unit) dedup

    for unit, path in CODEBOOK_FILES.items():
        if not path.exists():
            continue
        for line in path.read_text().splitlines():
            if not line.startswith("| `"):
                continue
            # Line format: | `VARNAME` | Label text | [preselected]? |
            parts = [p.strip() for p in line.split("|")]
            if len(parts) < 4:
                continue
            varname = parts[1].strip("`")
            label = parts[2]
            notes = parts[3]
            if pattern.search(varname) or pattern.search(label):
                key = (varname, unit)
                if key not in seen:
                    seen.add(key)
                    results.append({
                        "variable": varname,
                        "label": label,
                        "unit": unit,
                        "preselected": "[preselected]" in notes,
                    })

    results.sort(key=lambda r: (r["unit"], r["variable"]))
    return results


# ---------------------------------------------------------------------------
# submit_extract
# ---------------------------------------------------------------------------


def submit_extract(
    samples: list[str],
    variables: list[str],
    description: str = "ipums_dhs.py extract",
) -> tuple[int, str]:
    """
    Submit a DHS extract via the raw HTTP workaround (ipumspy's submit_extract
    does not work for DHS as of v0.7.0 — see CLAUDE.md).

    Returns:
        (extract_id, status)
    """
    api_key = _require_api_key()
    payload = {
        "description": description,
        "dataFormat": "csv",
        "dataStructure": {"rectangular": {"on": "P"}},
        "samples": {s: {} for s in samples},
        "variables": {v: {} for v in variables},
        "collection": "dhs",
        "version": 2,
    }
    resp = requests.post(
        IPUMS_API,
        params=IPUMS_PARAMS,
        json=payload,
        headers={"Authorization": api_key},
    )
    if not resp.ok:
        detail = resp.json().get("detail", resp.text)
        raise RuntimeError(f"Extract submission failed ({resp.status_code}): {detail}")
    data = resp.json()
    return data["number"], data["status"]


# ---------------------------------------------------------------------------
# wait_for_extract
# ---------------------------------------------------------------------------


def wait_for_extract(
    extract_id: int,
    poll_interval: int = 30,
    timeout: int = 3600,
    verbose: bool = True,
) -> dict:
    """
    Poll until the extract is completed (or failed). Returns the final API response dict.

    Raises RuntimeError on failure or timeout.
    """
    api_key = _require_api_key()
    elapsed = 0
    while elapsed < timeout:
        resp = requests.get(
            f"{IPUMS_API}/{extract_id}",
            params=IPUMS_PARAMS,
            headers={"Authorization": api_key},
        )
        resp.raise_for_status()
        data = resp.json()
        status = data["status"]
        if verbose:
            print(f"  [{elapsed}s] extract {extract_id} status={status}")
        if status == "completed":
            return data
        if status in ("failed", "cancelled"):
            raise RuntimeError(f"Extract {extract_id} {status}: {data}")
        time.sleep(poll_interval)
        elapsed += poll_interval
    raise RuntimeError(f"Extract {extract_id} timed out after {timeout}s")


# ---------------------------------------------------------------------------
# download_extract
# ---------------------------------------------------------------------------


def download_extract(extract_id: int, download_links: dict) -> Path:
    """
    Download the CSV data file and DDI XML codebook for an extract.

    Saves:
      data/extract_{id}.csv.gz  — the microdata
      data/extract_{id}.xml     — the DDI codebook (used for auto missing-value detection)

    Returns the path to the CSV file.
    """
    api_key = _require_api_key()
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    for key, suffix in [("data", ".csv.gz"), ("ddiCodebook", ".xml")]:
        url = download_links[key]["url"]
        resp = requests.get(url, headers={"Authorization": api_key})
        resp.raise_for_status()
        dest = DATA_DIR / f"extract_{extract_id}{suffix}"
        dest.write_bytes(resp.content)

    return DATA_DIR / f"extract_{extract_id}.csv.gz"


# ---------------------------------------------------------------------------
# load_extract
# ---------------------------------------------------------------------------


def load_extract(filepath: Path) -> pd.DataFrame:
    """Load a .csv.gz extract file into a pandas DataFrame."""
    with gzip.open(filepath, "rt") as f:
        return pd.read_csv(f)


# ---------------------------------------------------------------------------
# DDI missing-value detection
# ---------------------------------------------------------------------------

# Label substrings (lowercase) that identify a value code as missing/NIU.
_MISSING_LABELS = (
    "niu",
    "not in universe",
    "missing",
    "flagged",
    "don't know",
    "dont know",
    "unknown",
    "plausible limits",  # "out of plausible limits"
    "inconsistent",
)


def parse_ddi_missing(ddi_path: Path) -> dict[str, set]:
    """
    Parse a DDI XML codebook and return missing/NIU value codes per variable.

    Returns:
        Dict mapping variable name → set of integer values that represent
        missing, NIU, flagged, or otherwise invalid responses.

    Uses ipumspy's readers.read_ipums_ddi() to parse the DDI. The `codes`
    attribute of each VariableDescription is {label: value}; we flag any
    code whose label matches a known missing-value pattern.
    """
    ddi = readers.read_ipums_ddi(str(ddi_path))
    missing_codes: dict[str, set] = {}

    for var in ddi.data_description:
        bad = set()
        for label, value in var.codes.items():
            if any(pat in label.lower() for pat in _MISSING_LABELS):
                bad.add(value)
        if bad:
            missing_codes[var.name] = bad

    return missing_codes


def get_ddi_var_info(ddi_path: Path, variable: str) -> tuple[str, dict]:
    """
    Return (label, {value: label_str}) for a variable from a DDI codebook file.
    Returns ("", {}) if the variable is not found or the DDI cannot be read.
    """
    try:
        ddi = readers.read_ipums_ddi(str(ddi_path))
        for var_desc in ddi.data_description:
            if var_desc.name == variable:
                return var_desc.label, {v: lbl for lbl, v in var_desc.codes.items()}
    except Exception:
        pass
    return "", {}


def detect_categorical(series: pd.Series, value_labels: dict) -> bool:
    """
    True if more than half of the unique non-missing values in series have DDI labels.
    Signals that the variable is categorical and should show a full frequency table.
    """
    unique_vals = set(series.dropna().unique())
    if not unique_vals or not value_labels:
        return False
    labeled = {v for v in unique_vals if v in value_labels}
    # 50% threshold: a variable like WEALTHQ has 5 values all labeled (100%),
    # while a continuous variable like HWHAZWHO has hundreds of unique integer
    # z-score values only a handful of which (sentinel codes) are labeled.
    # The threshold is intentionally loose so that semi-continuous variables with
    # a few unlabeled outlier codes still read as categorical.
    return len(labeled) / len(unique_vals) > 0.5


def weighted_freq_table(
    df: pd.DataFrame,
    variable: str,
    weight_col: str,
    value_labels: dict,
    by: str = None,
) -> pd.DataFrame:
    """
    Compute a weighted frequency table for a categorical variable.

    Returns a DataFrame with columns: [by (optional), value, label, proportion, n].
    """
    def _freq(sub: pd.DataFrame) -> pd.DataFrame:
        valid = sub[sub[variable].notna()]
        total_w = valid[weight_col].sum()
        rows = []
        for val in sorted(valid[variable].unique()):
            mask = valid[variable] == val
            w = (valid[weight_col] * mask).sum()
            rows.append({
                "value": val,
                "label": value_labels.get(val, str(val)),
                "proportion": w / total_w if total_w > 0 else float("nan"),
                "n": int(mask.sum()),
            })
        return pd.DataFrame(rows)

    if by is None:
        return _freq(df)

    groups = []
    for grp_val in sorted(df[by].dropna().unique()):
        freq = _freq(df[df[by] == grp_val])
        freq.insert(0, by, grp_val)
        groups.append(freq)

    overall = _freq(df)
    overall.insert(0, by, "Overall")
    groups.append(overall)

    return pd.concat(groups, ignore_index=True)


# ---------------------------------------------------------------------------
# Weighted statistics
#
# DHS anthropometric z-scores (e.g. HWHAZWHO — height-for-age) are stored in
# IPUMS as integers multiplied by 100. A child with a HAZ of -1.73 is stored as
# -173. This avoids floating-point storage in the original survey data files.
# detect_scale() sniffs for this pattern from the DDI label and value range;
# apply_missing_and_scale() then divides by 100 so downstream statistics are in
# proper z-score units (e.g. cutoff of -2 SD for stunting works correctly).
# ---------------------------------------------------------------------------


def detect_scale(
    series: pd.Series,
    ddi_label: str = "",
) -> Optional[float]:
    """
    Heuristically detect whether a numeric variable is stored at an integer
    scale that should be divided out before analysis.

    Currently detects one case:
      - Z-score variables stored as integer × 100
        Trigger: DDI label contains "standard deviations" AND valid values
        fall in the range [-3000, 3000] with no fractional part, suggesting
        they represent hundredths rather than whole units.

    Returns the scale divisor (e.g. 100.0) if detected, else None.
    """
    if "standard deviations" not in ddi_label.lower():
        return None

    valid = series.dropna()
    if valid.empty:
        return None

    # Values stored ×100 will be integers in roughly [-600, 600].
    # Real z-scores are in [-6, 6]. Threshold: abs(median) > 10 is a
    # strong signal that the series is not already in z-score units.
    if valid.abs().median() > 10 and valid.abs().max() <= 3000:
        return 100.0

    return None


def apply_missing_and_scale(
    df: pd.DataFrame,
    variable: str,
    missing_ge: float = None,
    scale: float = None,
) -> pd.DataFrame:
    """
    Return a copy of df with sentinel values removed and/or the variable rescaled.

    Many DHS variables use high codes (9, 99, 999, 9995-9999, etc.) for
    "not in universe" or "don't know". Prefer passing a set of exact values
    via `missing_values` (derived from the DDI codebook via parse_ddi_missing)
    over the blunt `missing_ge` threshold. Both can be combined; DDI values
    are applied first, then the threshold.

    Some variables (e.g. HWHAZWHO, anthropometric z-scores) are stored as
    integers scaled by 100. Pass scale=100 to divide the column by 100.
    """
    df = df.copy()
    if missing_ge is not None:
        df.loc[df[variable] >= missing_ge, variable] = float("nan")
    if scale is not None and scale != 1:
        df[variable] = df[variable] / scale
    return df


def weighted_proportion(
    df: pd.DataFrame,
    variable: str,
    value,
    weight_col: str,
) -> float:
    """
    Compute the weighted proportion of rows where variable == value.
    Rows where variable is NaN are excluded.
    """
    mask = df[variable].notna()
    sub = df[mask]
    if sub.empty or sub[weight_col].sum() == 0:
        return float("nan")
    return (sub[weight_col] * (sub[variable] == value)).sum() / sub[weight_col].sum()


def weighted_mean(
    df: pd.DataFrame,
    variable: str,
    weight_col: str,
) -> float:
    """Compute the weighted mean of a numeric variable. NaN rows are excluded."""
    mask = df[variable].notna()
    sub = df[mask]
    if sub.empty or sub[weight_col].sum() == 0:
        return float("nan")
    return (sub[weight_col] * sub[variable]).sum() / sub[weight_col].sum()


def weighted_median(
    df: pd.DataFrame,
    variable: str,
    weight_col: str,
) -> float:
    """
    Compute the weighted median of a numeric variable. NaN rows are excluded.

    Algorithm: sort by value, compute cumulative weights, interpolate linearly
    at the point where cumulative weight crosses 50% of total weight.
    """
    mask = df[variable].notna()
    sub = df[mask].sort_values(variable)
    if sub.empty or sub[weight_col].sum() == 0:
        return float("nan")
    vals = sub[variable].values
    weights = sub[weight_col].values
    cumw = weights.cumsum()
    half = weights.sum() / 2
    idx = int((cumw >= half).argmax())
    if idx == 0:
        return float(vals[0])
    v1, v2 = float(vals[idx - 1]), float(vals[idx])
    w1, w2 = float(cumw[idx - 1]), float(cumw[idx])
    if w2 == w1:
        return float(v2)
    return v1 + (half - w1) / (w2 - w1) * (v2 - v1)


def tabulate(
    df: pd.DataFrame,
    variable: str,
    weight_col: str,
    by: str = None,
    below: float = None,
    median: bool = False,
) -> pd.DataFrame:
    """
    Compute weighted statistics for `variable`, optionally grouped by `by`.

    - `below`: proportion of valid rows where variable < below (e.g. -2 for stunting)
    - `median`: weighted median instead of mean
    - neither: weighted mean

    Returns a DataFrame with columns [group/by, stat, n]; stat column name encodes
    the statistic used.
    """
    if below is not None:
        stat_col = f"pct_below_{below}"

        def _stat(sub: pd.DataFrame) -> float:
            valid = sub[sub[variable].notna()]
            if valid.empty or valid[weight_col].sum() == 0:
                return float("nan")
            return (valid[weight_col] * (valid[variable] < below)).sum() / valid[weight_col].sum()

    elif median:
        stat_col = "weighted_median"

        def _stat(sub: pd.DataFrame) -> float:
            return weighted_median(sub, variable, weight_col)

    else:
        stat_col = "weighted_mean"

        def _stat(sub: pd.DataFrame) -> float:
            return weighted_mean(sub, variable, weight_col)

    def _row(sub: pd.DataFrame) -> dict:
        valid = sub[sub[variable].notna()]
        return {stat_col: _stat(sub), "n": len(valid)}

    if by is None:
        return pd.DataFrame([{"group": "Overall", **_row(df)}])

    rows = []
    for val in sorted(df[by].dropna().unique()):
        rows.append({by: val, **_row(df[df[by] == val])})
    rows.append({by: "Overall", **_row(df)})
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# CLI helpers
# ---------------------------------------------------------------------------


def load_availability() -> dict:
    """Load references/dhs_availability.json if it exists, else return empty dict."""
    if AVAILABILITY_PATH.exists():
        return json.loads(AVAILABILITY_PATH.read_text())
    return {}


def _best_sample_from_availability(
    country: str,
    unit: str,
    variables: list[str],
    all_unit_samples: list[dict],
    availability: dict,
) -> Optional[list[str]]:
    """
    Use pre-scraped availability data to find the newest sample that has all variables.

    Returns the list of sample_ids for the best year, or None when:
    - any variable is missing from the availability JSON
    - the country isn't found in a variable's availability entry
    - no IPUMS sample year aligns with the availability years

    Allows ±1 year tolerance because DHS survey years (stored in the availability
    JSON) can differ by one year from IPUMS sample-ID years (which reflect the
    data-release year, e.g., Malawi 2016 survey → IPUMS sample mw2017kr).
    """
    country_name = COUNTRY_NAMES.get(country.lower())
    if not country_name:
        return None

    # For each variable, collect the DHS survey years available for this country
    var_avail_years: dict[str, list[int]] = {}
    for var in variables:
        var_data = availability.get(var)
        if var_data is None:
            return None  # variable not scraped yet
        # Handle both old format {country: [years]} and new format {"availability": {...}, ...}
        avail_dict = var_data.get("availability", var_data) if isinstance(var_data.get("availability"), dict) else var_data
        country_years = avail_dict.get(country_name)
        if country_years is None:
            return None  # variable not available in this country
        var_avail_years[var] = country_years

    # Build year → sample_ids from the already-fetched samples list
    by_year: dict[int, list[str]] = {}
    for s in all_unit_samples:
        by_year.setdefault(s["year"], []).append(s["sample_id"])

    def _year_matches(ipums_year: int, dhs_years: list[int]) -> bool:
        # ±1 year tolerance: DHS fieldwork often spans a calendar year boundary,
        # and IPUMS assigns the data-release year rather than the fieldwork year.
        # Example: the Malawi 2015–16 DHS is labelled 2016 in the availability
        # JSON (scraped from IPUMS variable pages) but released as mw2017kr in
        # the IPUMS sample catalogue. Without this tolerance the lookup fails and
        # falls back to an unnecessary API brute-force search.
        return any(abs(ipums_year - y) <= 1 for y in dhs_years)

    # Walk IPUMS sample years newest → oldest; return first that covers all vars
    for ipums_year in sorted(by_year.keys(), reverse=True):
        if all(_year_matches(ipums_year, var_avail_years[v]) for v in variables):
            return by_year[ipums_year]

    return None


def _require_api_key() -> str:
    key = os.environ.get("IPUMS_API_KEY")
    if not key:
        sys.exit("Error: IPUMS_API_KEY environment variable not set.")
    return key


def _is_unavailable_error(err: RuntimeError) -> bool:
    """Return True if the error is a variable-availability rejection from the API."""
    return "not available in any of the samples" in str(err)


def _unavailable_vars(err: RuntimeError) -> list[str]:
    """Extract variable names from a variable-availability error message."""
    # Error text looks like: "['HWHAZWHO: This variable is not available...']"
    return re.findall(r"\b([A-Z][A-Z0-9_]+)\b(?=: This variable is not available)", str(err))


def _resolve_samples(country: str, survey: str, unit: str) -> list[str]:
    """
    Return a list of sample IDs matching country + survey + unit.

    survey = "latest"  → most recent year for that country/unit
    survey = "all"     → every available year
    survey = "YYYY"    → specific year (e.g. "2014")
    survey = explicit sample ID (e.g. "ke2014kr") → use as-is
    """
    suffix = UNIT_SUFFIX.get(unit)
    if not suffix:
        sys.exit(f"Unknown unit: {unit!r}. Choose from: {', '.join(UNIT_SUFFIX)}")

    # If it looks like an explicit sample ID, use it directly
    if re.match(r"^[a-z]{2}\d{4}[a-z]{2}$", survey.lower()):
        return [survey.lower()]

    all_samples = find_samples(country, unit)
    if not all_samples:
        sys.exit(f"No {unit} samples found for country {country!r}.")

    if survey == "latest":
        latest_year = max(s["year"] for s in all_samples)
        return [s["sample_id"] for s in all_samples if s["year"] == latest_year]

    if survey == "all":
        return [s["sample_id"] for s in all_samples]

    # Try matching as a year
    if re.match(r"^\d{4}$", survey):
        matches = [s["sample_id"] for s in all_samples if str(s["year"]) == survey]
        if not matches:
            available = sorted({str(s["year"]) for s in all_samples})
            sys.exit(f"No {unit} sample for {country} {survey}. Available: {', '.join(available)}")
        return matches

    sys.exit(f"Cannot interpret --survey {survey!r}. Use 'latest', 'all', a year, or a sample ID.")


def _print_table(
    result: pd.DataFrame,
    variable: str,
    by: str = None,
    by_labels: dict = None,
) -> None:
    def _fmt_group(val) -> str:
        if by_labels and val != "Overall":
            try:
                code = int(float(val))
                label = by_labels.get(code)
                if label:
                    return f"{label} ({code})"
            except (ValueError, TypeError):
                pass
        return str(val)

    col_width = max(len(_fmt_group(v)) for v in result.iloc[:, 0]) + 2
    header_var = by if by else "Group"
    stat_col = result.columns[1]
    if stat_col.startswith("pct_below"):
        header_stat, fmt = "% below cutoff", "{:.1%}"
    elif stat_col == "weighted_median":
        header_stat, fmt = "Weighted median", "{:.4f}"
    else:
        header_stat, fmt = "Weighted mean", "{:.4f}"

    print(f"\n{'─' * 60}")
    print(f"  {header_var:<{col_width}}  {header_stat:>14}  {'N':>8}")
    print(f"{'─' * 60}")
    for _, row in result.iterrows():
        group = _fmt_group(row.iloc[0])
        stat = row[stat_col]
        n = int(row["n"])
        stat_str = fmt.format(stat) if pd.notna(stat) else "  N/A"
        print(f"  {group:<{col_width}}  {stat_str:>14}  {n:>8,}")
    print(f"{'─' * 60}\n")


def _print_freq_table(result: pd.DataFrame, variable: str, by: str = None) -> None:
    """Print a weighted frequency table for a categorical variable."""
    def _display(row) -> str:
        return f"{row['label']} ({int(row['value'])})"

    label_w = max((len(_display(r)) for _, r in result.iterrows()), default=10)
    label_w = max(label_w, 10)
    width = label_w + 28

    def _rows(sub: pd.DataFrame) -> None:
        print(f"  {'─' * width}")
        print(f"  {'Label':<{label_w}}  {'%':>8}  {'N':>8}")
        print(f"  {'─' * width}")
        for _, row in sub.iterrows():
            pct = f"{row['proportion']:.1%}" if pd.notna(row["proportion"]) else "  N/A"
            print(f"  {_display(row):<{label_w}}  {pct:>8}  {int(row['n']):>8,}")
        print(f"  {'─' * width}")

    if by is None:
        _rows(result)
    else:
        all_vals = list(result[by].unique())
        numeric_vals = sorted(v for v in all_vals if v != "Overall")
        ordered = numeric_vals + (["Overall"] if "Overall" in all_vals else [])
        for grp_val in ordered:
            sub = result[result[by] == grp_val]
            heading = "Overall" if grp_val == "Overall" else f"{by} = {int(grp_val)}"
            print(f"\n  {heading}")
            _rows(sub)
    print()


# ---------------------------------------------------------------------------
# CLI commands
# ---------------------------------------------------------------------------


def cmd_samples(args: argparse.Namespace) -> None:
    samples = find_samples(args.country_code)
    if not samples:
        print(f"No samples found for country code {args.country_code!r}.")
        return

    # Group by unit for readable output
    from itertools import groupby
    by_unit = {}
    for s in samples:
        by_unit.setdefault(s["unit"], []).append(s)

    print(f"\nDHS samples for {args.country_code.upper()}:\n")
    for unit in UNIT_SUFFIX:  # preserve canonical order
        if unit not in by_unit:
            continue
        print(f"  {unit}:")
        for s in by_unit[unit]:
            print(f"    {s['sample_id']}  {s['description']}")
    print()


def cmd_search(args: argparse.Namespace) -> None:
    results = find_variables(args.keyword)
    if not results:
        print(f"No variables found matching {args.keyword!r}.")
        return

    print(f"\nVariables matching {args.keyword!r}:\n")
    print(f"  {'Variable':<25} {'Unit':<20} {'Label'}")
    print(f"  {'─'*25} {'─'*20} {'─'*40}")
    for r in results:
        pre = " *" if r["preselected"] else ""
        print(f"  {r['variable']:<25} {r['unit']:<20} {r['label']}{pre}")
    print(f"\n  * = preselected (auto-included in every extract)\n")


def cmd_table(args: argparse.Namespace) -> None:
    # Overall flow:
    #  1. Resolve which sample(s) to use, using pre-scraped availability data when
    #     possible so we don't submit blind extract requests for the wrong year.
    #  2. Submit an extract via raw HTTP (ipumspy's submit_extract doesn't work for
    #     DHS — see CLAUDE.md). Retry on older surveys if a variable is unavailable.
    #  3. Wait for the extract, download data (.csv.gz) and codebook (.xml / DDI).
    #  4. Apply row filter (--filter), then scrub missing/NIU values using DDI codes.
    #  5. Auto-detect scale (z-scores × 100) and categorical vs. continuous.
    #  6. Compute weighted statistics and print the table(s).
    #  7. Optionally write an XLSX file with formatted tables + replication block.
    if args.median and args.below is not None:
        sys.exit("--median and --below are mutually exclusive.")

    unit = args.unit
    weight_col = args.weight or WEIGHT_DEFAULT.get(unit, "PERWEIGHT")
    variables_requested = [v.strip() for v in args.variables.split(",")]

    # Parse --filter specs. Each spec is VARIABLE[>=<=]VALUE; multiple are ANDed together.
    _FILTER_RE = re.compile(r"^([A-Za-z_]\w*)(>=|<=|>|<|=)(-?[\d.]+)$")
    _FILTER_OPS = {
        "==": lambda s, v: s == v,
        ">=": lambda s, v: s >= v,
        "<=": lambda s, v: s <= v,
        ">":  lambda s, v: s > v,
        "<":  lambda s, v: s < v,
    }

    def _parse_filter(spec: str) -> dict:
        m = _FILTER_RE.match(spec.strip())
        if not m:
            sys.exit(
                f"--filter must be in VARIABLE[>=<=]VALUE form (e.g. KIDALIVE=1 or "
                f"KIDCURAGEMO>=12), got: {spec!r}"
            )
        var, op, val_raw = m.group(1).upper(), m.group(2), m.group(3)
        op = "==" if op == "=" else op  # normalise single = to ==
        try:
            val: float | int = int(val_raw)
        except ValueError:
            val = float(val_raw)
        return {"var": var, "op": op, "val": val}

    filters = [_parse_filter(s) for s in (args.filter_specs or [])]
    filter_vars = list(dict.fromkeys(f["var"] for f in filters))

    # Build full variable list for extract: weight + by variable + all filter variables
    extract_vars = list(dict.fromkeys(
        variables_requested + [weight_col]
        + ([args.by] if args.by else [])
        + filter_vars
    ))

    # 1. Build ordered candidate list (newest first) for fallback retries.
    #    For explicit sample IDs or "all", no fallback is attempted.
    print(f"\nResolving samples for {args.country.upper()} {unit} ({args.survey})...")
    all_unit_samples = find_samples(args.country, unit)  # sorted by year asc
    country_name = all_unit_samples[0]["country"] if all_unit_samples else args.country.upper()
    candidates: list[list[str]] = []

    if args.survey == "latest":
        # Use pre-scraped availability data if present — jump directly to the
        # newest survey year that actually contains all requested variables,
        # rather than submitting extract requests and parsing API error messages.
        # Falls back to newest-to-oldest iteration when a variable isn't in the
        # availability JSON (e.g. newly added variables not yet scraped).
        by_year: dict[int, list[str]] = {}
        for s in all_unit_samples:
            by_year.setdefault(s["year"], []).append(s["sample_id"])

        availability = load_availability()
        if availability:
            best = _best_sample_from_availability(
                args.country, unit, variables_requested, all_unit_samples, availability
            )
            if best:
                print(f"  Availability data: {', '.join(best)}")
                candidates = [best]
            else:
                print("  Availability data found but variable/country not covered — "
                      "falling back to newest-to-oldest search")
                for year in sorted(by_year, reverse=True):
                    candidates.append(by_year[year])
        else:
            # No availability data — try each year from newest to oldest
            for year in sorted(by_year, reverse=True):
                candidates.append(by_year[year])
    else:
        # Non-latest surveys: single attempt, no fallback
        candidates = [_resolve_samples(args.country, args.survey, unit)]

    if not candidates:
        sys.exit(f"No {unit} samples found for country {args.country!r}.")

    # Submit extract with fallback: if the API rejects because a variable isn't in
    # the chosen survey year, pop it off the candidates list and retry the next
    # older year. This handles the case where availability data is stale or absent.
    # We use raw HTTP rather than ipumspy.submit_extract() because the DHS API
    # rejects the attachedCharacteristics field that ipumspy always sends — a DHS
    # API limitation not present in other IPUMS collections (see CLAUDE.md).
    final = filepath = None
    sample_ids = None
    for attempt, sample_ids in enumerate(candidates):
        print(f"  Trying: {', '.join(sample_ids)}")
        try:
            print(f"Submitting extract (variables: {', '.join(extract_vars)})...")
            extract_id, status = submit_extract(
                sample_ids,
                extract_vars,
                description=f"{args.country.upper()} {unit} — {', '.join(variables_requested)}",
            )
            print(f"  Extract ID: {extract_id}  initial status: {status}")

            print("Waiting for extract to complete...")
            final = wait_for_extract(extract_id, verbose=True)

            print("Downloading...")
            filepath = download_extract(extract_id, final["downloadLinks"])
            ddi_path = filepath.with_suffix("").with_suffix(".xml")
            print(f"  Saved to {filepath}")
            print(f"  DDI at  {ddi_path}")
            break  # success

        except RuntimeError as e:
            if not _is_unavailable_error(e):
                raise
            bad_vars = _unavailable_vars(e) or variables_requested
            remaining = candidates[attempt + 1:]
            if remaining:
                next_ids = remaining[0]
                print(f"  {', '.join(bad_vars)} not available in "
                      f"{', '.join(sample_ids)}. Trying {', '.join(next_ids)}...")
            else:
                available = [s["sample_id"] for s in all_unit_samples]
                sys.exit(
                    f"\nNo {unit} sample for {args.country.upper()} has all requested variables.\n"
                    f"Available samples: {', '.join(available)}\n"
                    f"Check variable availability at: "
                    f"https://www.idhsdata.org/idhs-action/variables/{bad_vars[0]}"
                )

    if final is None:
        sys.exit("No extract was completed.")

    # 5. Load
    df = load_extract(filepath)
    print(f"  Loaded {len(df):,} rows × {len(df.columns)} columns")

    # 5a. Apply --filter(s) — each filter is ANDed in sequence
    for f in filters:
        var, op, val = f["var"], f["op"], f["val"]
        if var not in df.columns:
            sys.exit(f"Filter variable {var!r} not found in extract columns.")
        df = df[_FILTER_OPS[op](df[var], val)].copy()
        print(f"  Applied filter: {var} {op} {val} ({len(df):,} rows remaining)")

    # Parse the DDI XML codebook that IPUMS ships alongside every extract.
    # It contains the exact sentinel codes (NIU, "don't know", flagged, etc.)
    # for each variable. Using DDI codes is much more reliable than a blanket
    # threshold like >= 9000 — different variables use different sentinel ranges,
    # and a threshold would silently drop valid high values (e.g. a woman aged 49).
    ddi_missing: dict[str, set] = {}
    if not args.no_ddi_filter:
        ddi_missing = parse_ddi_missing(ddi_path)
        if ddi_missing:
            flagged = {v: sorted(c) for v, c in ddi_missing.items() if v in extract_vars}
            if flagged:
                print(f"  DDI missing codes: { {v: c for v, c in flagged.items()} }")

    # 5c. Pre-load DDI value labels for --by variable (used in _print_table for labelled rows)
    by_value_labels: dict = {}
    if args.by and not args.no_ddi_filter:
        _, by_raw = get_ddi_var_info(ddi_path, args.by)
        by_missing = ddi_missing.get(args.by, set())
        by_value_labels = {v: lbl for v, lbl in by_raw.items() if v not in by_missing}

    # 6. Compute and print results
    replication_vars: dict = {}
    missing_stats: dict = {}  # {var: {total, rows_after, by_code: {code: {label, count}}}}
    xlsx_results: list = []  # [(var, result_df), ...] collected for XLSX output
    for var in variables_requested:
        if var not in df.columns:
            print(f"  Warning: {var!r} not in extract columns. Skipping.")
            continue
        if var == weight_col or var == args.by:
            continue

        # Apply DDI-detected missing values first, then manual --missing-ge override
        working = df.copy()
        ddi_vals = ddi_missing.get(var, set())
        if ddi_vals:
            n_before = working[var].notna().sum()
            working.loc[working[var].isin(ddi_vals), var] = float("nan")
            n_dropped_ddi = n_before - working[var].notna().sum()
            print(f"  {var}: filtered {n_dropped_ddi:,} DDI missing rows "
                  f"(codes: {sorted(ddi_vals)})")

        if args.by and not args.no_ddi_filter:
            _by_missing = ddi_missing.get(args.by, set())
            if _by_missing:
                working = working[~working[args.by].isin(_by_missing)]

        if args.missing_ge is not None:
            n_before = working[var].notna().sum()
            working.loc[working[var] >= args.missing_ge, var] = float("nan")
            n_dropped_ge = n_before - working[var].notna().sum()
            if n_dropped_ge:
                print(f"  {var}: filtered {n_dropped_ge:,} additional rows (>= {args.missing_ge})")

        # Get DDI info: variable label (for scale detection) and value codes (for categorical detection)
        ddi_var_label, raw_value_labels = "", {}
        if not args.no_ddi_filter:
            ddi_var_label, raw_value_labels = get_ddi_var_info(ddi_path, var)

        # Exclude missing codes from value labels so they don't influence categorical detection
        ddi_vals_for_var = ddi_missing.get(var, set())
        value_labels = {v: lbl for v, lbl in raw_value_labels.items()
                        if v not in ddi_vals_for_var}

        # Auto-detect scale from DDI label + value distribution (unless overridden)
        scale = args.scale
        if scale is None and not args.no_ddi_filter:
            scale = detect_scale(working[var], ddi_var_label)
            if scale is not None:
                print(f"  {var}: auto-detected scale={scale} "
                      f"(label contains 'standard deviations', values look like integers × {int(scale)})")

        if scale and scale != 1:
            working[var] = working[var] / scale
            if args.scale:  # only print if manually specified (auto-detect already printed above)
                print(f"  {var}: values divided by {scale}")

        replication_vars[var] = {
            "label": ddi_var_label,
            "raw_codes": raw_value_labels,
            "missing_codes": sorted(ddi_missing.get(var, set())),
            "scale": scale if (scale and scale != 1) else None,
            "below": args.below,
        }

        # Count missing code occurrences against the original (unfiltered) df so the
        # XLSX missing-values summary reflects the full survey, not just the analysis
        # subset (e.g. after --filter HHLINENO=1 the filtered df would undercount NIU rows).
        _code_details: dict = {}
        for code in sorted(ddi_missing.get(var, set())):
            _code_details[code] = {
                "label": raw_value_labels.get(code, ""),
                "count": int((df[var] == code).sum()),
            }
        missing_stats[var] = {
            "total": len(df),
            "rows_after": int(working[var].notna().sum()),
            "by_code": _code_details,
        }

        # Detect categorical vs continuous. Force continuous if --below or --median is specified.
        is_cat = args.below is None and not args.median and detect_categorical(working[var], value_labels)

        print(f"\nVariable: {var}")

        if is_cat:
            # Full weighted frequency table with DDI labels per value
            if args.survey == "all":
                years = sorted(working["YEAR"].dropna().unique().astype(int))
                if args.by is None:
                    result = weighted_freq_table(working, var, weight_col,
                                                 value_labels, by="YEAR")
                    _print_freq_table(result, var, by="YEAR")
                else:
                    all_results = []
                    for year in years:
                        year_df = working[working["YEAR"] == year]
                        result = weighted_freq_table(year_df, var, weight_col,
                                                     value_labels, by=args.by)
                        print(f"\n  ── {year} ──")
                        _print_freq_table(result, var, by=args.by)
                        result.insert(0, "YEAR", year)
                        all_results.append(result)
                    result = pd.concat(all_results, ignore_index=True)
            else:
                result = weighted_freq_table(working, var, weight_col,
                                             value_labels, by=args.by)
                _print_freq_table(result, var, by=args.by)
        else:
            # Continuous: weighted mean, or proportion below threshold
            if args.survey == "all":
                years = sorted(working["YEAR"].dropna().unique().astype(int))
                if args.by is None:
                    result = tabulate(working, var, weight_col, by="YEAR",
                                      below=args.below, median=args.median)
                    _print_table(result, var, by="YEAR")
                else:
                    all_results = []
                    for year in years:
                        year_df = working[working["YEAR"] == year]
                        result = tabulate(year_df, var, weight_col, by=args.by,
                                          below=args.below, median=args.median)
                        print(f"\n  ── {year} ──")
                        _print_table(result, var, by=args.by, by_labels=by_value_labels)
                        result.insert(0, "YEAR", year)
                        all_results.append(result)
                    result = pd.concat(all_results, ignore_index=True)
            else:
                result = tabulate(working, var, weight_col, by=args.by,
                                  below=args.below, median=args.median)
                _print_table(result, var, by=args.by, by_labels=by_value_labels)

        # 7. Collect result for XLSX output
        if args.output:
            xlsx_results.append((var, result))

        # 8. Optional: plot
        if args.plot:
            _plot(result, var, by=args.by, country=args.country, survey=args.survey)

    # 9. Replication block — collect DDI info for weight and by variables, then print
    for v in extract_vars:
        if v not in replication_vars:
            v_label, v_raw_codes = ("", {})
            if not args.no_ddi_filter and ddi_path:
                v_label, v_raw_codes = get_ddi_var_info(ddi_path, v)
            replication_vars[v] = {
                "label": v_label,
                "raw_codes": v_raw_codes,
                "missing_codes": sorted(ddi_missing.get(v, set())),
                "scale": None,
                "below": None,
            }
        if v not in missing_stats:
            v_raw_codes = replication_vars[v]["raw_codes"]
            v_ddi_vals = ddi_missing.get(v, set())
            _code_details = {}
            for code in sorted(v_ddi_vals):
                _code_details[code] = {
                    "label": v_raw_codes.get(code, ""),
                    "count": int((df[v] == code).sum()),
                }
            rows_after = int(df[v][~df[v].isin(v_ddi_vals)].notna().sum()) if v_ddi_vals else len(df)
            missing_stats[v] = {
                "total": len(df),
                "rows_after": rows_after,
                "by_code": _code_details,
            }

    # Build universe description for the replication block.  Precedence:
    #   1. --universe flag (explicit human-readable text from the caller)
    #   2. --filter(s) were applied: join all conditions
    #   3. Neither: state that no restrictions were applied
    if args.universe:
        universe_description = args.universe
    elif filters:
        parts = [f"{f['var']} {f['op']} {f['val']}" for f in filters]
        universe_description = " & ".join(parts)
    else:
        universe_description = (
            "No additional universe restrictions applied "
            "(full extract after DDI missing value filtering)"
        )

    _print_replication(
        sample_ids=sample_ids,
        country_name=country_name,
        unit=unit,
        weight_col=weight_col,
        extract_vars=extract_vars,
        variables_requested=variables_requested,
        replication_vars=replication_vars,
        median=args.median,
        universe_description=universe_description,
    )

    if args.output and xlsx_results:
        out_path = _write_xlsx_output(
            output_path=args.output,
            xlsx_results=xlsx_results,
            by=args.by,
            by_value_labels=by_value_labels,
            sample_ids=sample_ids,
            country_name=country_name,
            unit=unit,
            weight_col=weight_col,
            extract_vars=extract_vars,
            variables_requested=variables_requested,
            replication_vars=replication_vars,
            missing_stats=missing_stats,
            median=args.median,
            universe_description=universe_description,
        )
        print(f"  Saved to {out_path}")


def _plot(result: pd.DataFrame, variable: str, by: str, country: str, survey: str) -> None:
    try:
        import matplotlib.pyplot as plt
    except ImportError:
        print("  matplotlib not installed — skipping plot.")
        return

    plot_data = result[result.iloc[:, 0] != "Overall"].copy()
    x = plot_data.iloc[:, 0].astype(str)
    y = plot_data["weighted_mean"]

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(x, y)
    ax.set_xlabel(by or "Group")
    ax.set_ylabel(f"Weighted mean: {variable}")
    ax.set_title(f"{variable} by {by} — {country.upper()} ({survey})")
    plt.tight_layout()

    plot_path = DATA_DIR / f"{variable}_{by}_{country}_{survey}.png"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    fig.savefig(plot_path)
    print(f"  Plot saved to {plot_path}")
    plt.close(fig)


def _enrich_result_for_csv(
    result: pd.DataFrame, by: Optional[str], by_value_labels: dict
) -> pd.DataFrame:
    """Insert a {by}_label column after {by} for CSV export."""
    if not by or by not in result.columns or not by_value_labels:
        return result

    def _label(val) -> str:
        if val == "Overall":
            return "Overall"
        try:
            code = int(float(val))
            return by_value_labels.get(code, str(val))
        except (ValueError, TypeError):
            return str(val)

    result = result.copy()
    idx = result.columns.get_loc(by)
    result.insert(idx + 1, f"{by}_label", result[by].apply(_label))
    return result


def _write_xlsx_output(
    output_path: str,
    xlsx_results: list,
    by: Optional[str],
    by_value_labels: dict,
    sample_ids: list[str],
    country_name: str,
    unit: str,
    weight_col: str,
    extract_vars: list[str],
    variables_requested: list[str],
    replication_vars: dict,
    missing_stats: dict,
    filter_var: Optional[str] = None,
    filter_val=None,
    median: bool = False,
    universe_description: Optional[str] = None,
) -> str:
    """
    Write query results and provenance to a formatted Excel workbook.

    Sheet layout (top to bottom):
      1. Data table(s) — one per variable in variables_requested. Each table has a
         bold title row (merged across all data columns), a header row, and one data
         row per category/group. Percentage columns are formatted as 0.0%.
      2. Replication block — source, sample ID(s), weight, unit of analysis, filter
         (if any), per-variable IPUMS links, and any data transformations applied
         (z-score scaling, --below threshold).
      3. Comparability note (multi-year only) — italic warning that survey universes
         may differ across years.
      4. Missing values summary — total rows, rows excluded, rows after filtering for
         every variable in the extract.
      5. Missing values per-code detail — one row per sentinel code listing its label
         and raw count before any filtering.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter

    path = Path(output_path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_table(start_row, headers, data_rows, merge_value_to=None):
        r = start_row
        last_col = len(headers)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.border = bdr
            c.font = Font(bold=True)
        if merge_value_to and merge_value_to > last_col:
            ws.merge_cells(start_row=r, start_column=last_col, end_row=r, end_column=merge_value_to)
        r += 1
        for dr in data_rows:
            for ci, v in enumerate(dr, 1):
                ws.cell(row=r, column=ci, value=v).border = bdr
            if merge_value_to and merge_value_to > last_col:
                ws.merge_cells(start_row=r, start_column=last_col, end_row=r, end_column=merge_value_to)
            r += 1
        return r

    def _section_head(r, text):
        ws.cell(row=r, column=1, value=text).font = Font(bold=True, size=12)
        return r + 1

    unit_desc = UNIT_DESCRIPTIONS.get(unit, unit)
    if len(sample_ids) == 1:
        sid = sample_ids[0]
        year = sid[2:6]
        title_prefix = f"{country_name} {year} DHS"
        sample_line = f"{country_name} {year} DHS, {unit_desc} ({sid})"
    else:
        title_prefix = f"{country_name} DHS"
        sample_line = f"{country_name} DHS, {unit_desc} ({', '.join(sample_ids)})"

    by_ddi_label = ""
    if by:
        by_ddi_label = replication_vars.get(by, {}).get("label") or by

    n_data_cols = max(
        len(_enrich_result_for_csv(result, by, by_value_labels).columns)
        for _, result in xlsx_results
    )
    rep_end_col = max(n_data_cols, 4)
    measure_rows: set[int] = set()
    cur = 1

    # ── Data table(s) ──────────────────────────────────────────────────────────
    for var, result in xlsx_results:
        var_ddi_label = replication_vars.get(var, {}).get("label") or var

        # Title
        title = (
            f"{title_prefix} — {var_ddi_label} by {by_ddi_label}, {unit_desc}"
            if by_ddi_label
            else f"{title_prefix} — {var_ddi_label}, {unit_desc}"
        )
        _tc = ws.cell(row=cur, column=1, value=title)
        _tc.font = Font(bold=True, size=13)
        _tc.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[cur].height = 40
        if rep_end_col > 1:
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=rep_end_col)
        cur += 2

        enriched = _enrich_result_for_csv(result, by, by_value_labels)
        cols = list(enriched.columns)
        pct_cols = {c for c in cols if c == "proportion" or c.startswith("pct_below_")}

        def _header(col_name, _var=var, _vl=var_ddi_label, _bl=by_ddi_label):
            if col_name == by:
                return _bl or col_name
            if col_name == f"{by}_label":
                return "Label"
            if col_name == "value":
                return _vl
            if col_name == "label":
                return f"{_var} label"
            if col_name == "group":
                return "Group"
            if col_name == "proportion":
                return "%"
            if col_name == "n":
                return "N"
            if col_name == "YEAR":
                return "Year"
            if col_name.startswith("pct_below_"):
                threshold = col_name.replace("pct_below_", "")
                indicator = _INDICATOR_NAMES.get(_var, _vl)
                return f"{indicator} (% below {threshold})"
            if col_name == "weighted_mean":
                return f"{_vl} (weighted mean)"
            if col_name == "weighted_median":
                return f"{_vl} (weighted median)"
            return col_name

        # Header row
        data_section_start = cur
        for ci, col_name in enumerate(cols, 1):
            c = ws.cell(row=cur, column=ci, value=_header(col_name))
            c.border = bdr
            c.font = Font(bold=True)
        cur += 1

        # Data rows
        for _, row_data in enriched.iterrows():
            for ci, (col_name, val) in enumerate(zip(cols, row_data), 1):
                if isinstance(val, float) and pd.isna(val):
                    val = None
                elif col_name == "n":
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        pass
                elif col_name == "value" and val is not None:
                    try:
                        val = int(float(val))
                    except (ValueError, TypeError):
                        pass
                c = ws.cell(row=cur, column=ci, value=val)
                c.border = bdr
                if col_name in pct_cols and isinstance(val, float):
                    c.number_format = "0.0%"
            cur += 1
        measure_rows.update(range(data_section_start, cur))
        cur += 1  # blank row

    # ── Replication ────────────────────────────────────────────────────────────
    # Two-column key/value table. The "Value" cell is merged across all data
    # columns (up to rep_end_col) so long URLs don't bleed into adjacent cells.
    cur = _section_head(cur, "Replication")
    rep_rows = [
        ["Source", "IPUMS DHS (https://www.idhsdata.org)"],
        ["Sample", sample_line],
        ["Weight", weight_col],
        ["Unit of analysis", unit],
        ["Universe", universe_description],
    ]
    if median:
        rep_rows.append(["Statistic", "weighted median"])
    for v in extract_vars:
        rep_rows.append([v, f"https://www.idhsdata.org/idhs-action/variables/{v}"])
    for v in variables_requested:
        info = replication_vars.get(v, {})
        sc, bl = info.get("scale"), info.get("below")
        if sc:
            rep_rows.append(["Data transformation",
                              f"{v}: raw values stored as z-score × {int(sc)}, "
                              f"divided by {int(sc)} before analysis"])
        if bl is not None:
            indicator = _INDICATOR_NAMES.get(v, "Outcome")
            unit_str = " SD" if sc else ""
            rep_rows.append(["Data transformation",
                              f"{indicator}: defined as {v} < {bl}{unit_str}"])
    cur = _write_table(cur, ["Field", "Value"], rep_rows, merge_value_to=rep_end_col)

    # Comparability note: only shown when results span multiple survey years
    # (--survey all). DHS changed respondent universes over time — e.g. older
    # women's surveys covered ever-married women only — so trends need careful
    # interpretation.
    years = {sid[2:6] for sid in sample_ids}
    if len(years) > 1:
        note = (
            "Note: Older surveys may have different respondent universes and results may not be "
            "directly comparable across all years. See the Guide to DHS Statistics for details "
            "on changes over time for specific indicators."
        )
        nc = ws.cell(row=cur, column=1, value=note)
        nc.font = Font(italic=True)
        nc.alignment = Alignment(wrap_text=True)
        if rep_end_col > 1:
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=rep_end_col)
        ws.row_dimensions[cur].height = 40
        cur += 1

    cur += 1

    # ── Missing values summary ──────────────────────────────────────────────────
    # High-level row counts: how many rows were excluded per variable. Covers all
    # variables in the extract (including weight, by, filter) not just the analysis
    # variable, so reviewers can see the full NIU landscape.
    cur = _section_head(cur, "Missing values summary")
    mv_rows = [
        [v, s["total"], s["total"] - s["rows_after"], s["rows_after"]]
        for v, s in missing_stats.items()
    ]
    mv_start = cur
    cur = _write_table(
        cur,
        ["Variable", "Total rows", "Missing rows excluded", "Rows after filtering"],
        mv_rows,
    )
    measure_rows.update(range(mv_start, cur))
    cur += 1

    # ── Missing values per-code detail ──────────────────────────────────────────
    # One row per DDI sentinel code with its label (e.g. "NIU (not in universe)")
    # and raw count. Omitted entirely if no missing codes were detected.
    detail_rows = [
        [v, code, info["label"], info["count"]]
        for v, s in missing_stats.items()
        for code, info in s["by_code"].items()
    ]
    if detail_rows:
        detail_start = cur
        detail_end = _write_table(cur, ["Variable", "Missing code", "Label", "Count"], detail_rows)
        measure_rows.update(range(detail_start, detail_end))

    # ── Auto-fit column widths ──────────────────────────────────────────────────
    # Only measure cells in table rows (measure_rows); title and section-head rows
    # contain long strings that would make all columns far too wide.
    col_widths: dict[int, int] = {}
    for ws_row in ws.iter_rows():
        for cell in ws_row:
            if cell.value is not None and cell.row in measure_rows:
                col_widths[cell.column] = max(
                    col_widths.get(cell.column, 0), len(str(cell.value))
                )
    for col_idx, w in col_widths.items():
        max_w = 30 if col_idx == 1 else 20
        ws.column_dimensions[get_column_letter(col_idx)].width = min(w + 4, max_w)

    wb.save(path)
    return str(path)


def _print_replication(
    sample_ids: list[str],
    country_name: str,
    unit: str,
    weight_col: str,
    extract_vars: list[str],
    variables_requested: list[str],
    replication_vars: dict,
    filter_var: Optional[str] = None,
    filter_val=None,
    median: bool = False,
    universe_description: Optional[str] = None,
) -> None:
    """Print a replication citation and variable details block."""
    unit_desc = UNIT_DESCRIPTIONS.get(unit, unit)

    # Sample line
    if len(sample_ids) == 1:
        sid = sample_ids[0]
        year = sid[2:6]
        sample_line = f"{country_name} {year} DHS, {unit_desc} ({sid})"
    else:
        sample_line = f"{country_name} DHS, {unit_desc} ({', '.join(sample_ids)})"

    print("\n--- Replication ---")
    print(f"Source: IPUMS DHS (https://www.idhsdata.org)")
    print(f"Sample: {sample_line}")
    print(f"Weight: {weight_col}")
    print(f"Unit of analysis: {unit}")
    print(f"Universe: {universe_description}")
    if median:
        print("Statistic: weighted median")

    print("\nVariables:")
    for v in extract_vars:
        print(f"  {v} — https://www.idhsdata.org/idhs-action/variables/{v}")
    print("  See variable links above for full code definitions and value labels.")

    # Missing codes excluded
    has_missing = any(replication_vars.get(v, {}).get("missing_codes") for v in extract_vars)
    if has_missing:
        print("\nMissing codes excluded:")
        for v in extract_vars:
            codes = replication_vars.get(v, {}).get("missing_codes", [])
            if codes:
                print(f"  {v}: {', '.join(str(c) for c in codes)}")

    # Data transformations (only if scale or below applied)
    transforms = []
    for v in variables_requested:
        info = replication_vars.get(v, {})
        sc = info.get("scale")
        bl = info.get("below")
        if sc:
            transforms.append(
                f"  {v} raw values are stored as z-score × {int(sc)}"
                f" (divided by {int(sc)} before analysis)"
            )
        if bl is not None:
            indicator = _INDICATOR_NAMES.get(v, "Outcome")
            unit_str = " standard deviations" if sc else ""
            transforms.append(
                f"  {indicator} defined as {v} < {bl}{unit_str}"
            )
    if transforms:
        print("\nData transformations:")
        for t in transforms:
            print(t)



# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="ipums_dhs.py",
        description="Query IPUMS DHS microdata.",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # samples
    p_samples = sub.add_parser("samples", help="List DHS samples for a country.")
    p_samples.add_argument("country_code", help="Two-letter country code, e.g. KE")

    # search
    p_search = sub.add_parser("search", help="Search codebooks for a variable keyword.")
    p_search.add_argument("keyword", help="Search term (case-insensitive)")

    # table
    p_table = sub.add_parser("table", help="Fetch data and compute weighted statistics.")
    p_table.add_argument("--country", required=True, help="Two-letter country code")
    p_table.add_argument(
        "--survey",
        default="latest",
        help="'latest', 'all', a year (e.g. 2014), or an explicit sample ID",
    )
    p_table.add_argument("--variables", required=True, help="Comma-separated variable names")
    p_table.add_argument(
        "--unit",
        required=True,
        choices=list(UNIT_SUFFIX.keys()),
        help="Unit of analysis",
    )
    p_table.add_argument("--weight", default=None, help="Weight variable (default per unit)")
    p_table.add_argument("--by", default=None, help="Cross-tabulate by this variable")
    p_table.add_argument(
        "--filter",
        action="append",
        default=None,
        dest="filter_specs",
        metavar="VARIABLE[>=<=]VALUE",
        help=(
            "Filter rows before computing. Supports =, >=, <=, >, <. "
            "Repeatable: --filter KIDALIVE=1 --filter KIDCURAGEMO>=12 --filter KIDCURAGEMO<=23"
        ),
    )
    p_table.add_argument("--output", default=None, help="Save results to CSV file")
    p_table.add_argument("--plot", action="store_true", help="Generate a bar chart")
    p_table.add_argument(
        "--missing-ge",
        type=float,
        default=None,
        dest="missing_ge",
        help="Treat values >= N as missing (e.g. 9000 for DHS anthropometric sentinel codes)",
    )
    p_table.add_argument(
        "--scale",
        type=float,
        default=None,
        help="Divide variable values by N before computing (e.g. 100 for z-scores stored as integers)",
    )
    p_table.add_argument(
        "--below",
        type=float,
        default=None,
        help="Compute proportion of valid rows below this value instead of mean (e.g. -2 for stunting)",
    )
    p_table.add_argument(
        "--median",
        action="store_true",
        default=False,
        help="Compute weighted median instead of weighted mean (continuous variables only)",
    )
    p_table.add_argument(
        "--no-ddi-filter",
        action="store_true",
        dest="no_ddi_filter",
        help="Disable automatic DDI-based missing value filtering",
    )
    p_table.add_argument(
        "--universe",
        default=None,
        metavar="DESCRIPTION",
        help=(
            "Human-readable description of the universe/denominator for the replication block "
            "(e.g. 'Living children age 12-23 months'). "
            "If omitted, constructed from --filter when present, or states no restriction."
        ),
    )

    args = parser.parse_args()

    if args.command == "samples":
        cmd_samples(args)
    elif args.command == "search":
        cmd_search(args)
    elif args.command == "table":
        cmd_table(args)


if __name__ == "__main__":
    main()
